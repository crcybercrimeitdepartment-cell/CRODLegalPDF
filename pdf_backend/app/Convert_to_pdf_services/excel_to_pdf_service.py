import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, A3, A5, letter, legal, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

from app.core.paths import Paths

logger = logging.getLogger(__name__)

PAGE_SIZE_MAP = {
    "a4": A4,
    "a3": A3,
    "a5": A5,
    "letter": letter,
    "legal": legal,
}

MARGIN_PRESETS = {
    "normal": (0.75 * inch, 0.75 * inch, 1.0 * inch, 1.0 * inch),   # L, R, T, B
    "narrow": (0.25 * inch, 0.25 * inch, 0.75 * inch, 0.75 * inch),
    "wide":   (1.0 * inch,  1.0 * inch,  1.0 * inch,  1.0 * inch),
}


def _cell_value(cell) -> str:
    """Return a safe string representation of a cell value."""
    if cell.value is None:
        return ""
    return str(cell.value)


def _hex_to_rgb(hex_color: str):
    """Convert a hex color string (with or without #) to an RGB tuple (0-1 range)."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 6:
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        return r / 255, g / 255, b / 255
    return 0, 0, 0


def _sheet_to_table(ws, config: Dict[str, Any]) -> Table:
    """Convert an openpyxl worksheet into a reportlab Table."""
    show_gridlines = config.get("gridlines", False)
    show_headings = config.get("headings", False)

    # Collect all data rows
    data = []
    col_widths = []

    if show_headings:
        # Add column letter header row (A, B, C …)
        max_col = ws.max_column or 1
        header_row = [""] + [get_column_letter(c) for c in range(1, max_col + 1)]
        data.append(header_row)

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_data = []
        if show_headings:
            row_data.append(str(row_idx))
        for cell in row:
            row_data.append(_cell_value(cell))
        data.append(row_data)

    if not data:
        data = [["(empty sheet)"]]

    # Auto column widths (characters-based heuristic, capped)
    num_cols = max(len(r) for r in data)
    col_widths = []
    for col_idx in range(num_cols):
        max_len = max(
            len(str(row[col_idx])) if col_idx < len(row) else 0
            for row in data
        )
        width = min(max(max_len * 7, 40), 120)  # in points
        col_widths.append(width)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",(0, 0), (-1, -1), 3),
        ("TOPPADDING",  (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
        ("WORDWRAP",    (0, 0), (-1, -1), True),
    ]

    if show_gridlines:
        style_cmds += [
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ]
    else:
        style_cmds += [
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.black),
        ]

    # Apply header row bold style if headings are shown
    if show_headings:
        style_cmds += [
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
        ]

    # Apply cell fill colors from Excel where available
    for row_idx, row in enumerate(ws.iter_rows(), start=0):
        offset = 1 if show_headings else 0
        tbl_row = row_idx + offset
        for col_idx, cell in enumerate(row):
            t_col = col_idx + (1 if show_headings else 0)
            try:
                fill = cell.fill
                if fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.type == "rgb":
                    hex_val = fill.fgColor.rgb  # AARRGGBB
                    if hex_val and hex_val not in ("00000000", "FFFFFFFF", "FF000000"):
                        rgb = hex_val[2:]  # strip alpha
                        style_cmds.append(("BACKGROUND", (t_col, tbl_row), (t_col, tbl_row), colors.HexColor(f"#{rgb}")))
            except Exception:
                pass

            # Bold font
            try:
                if cell.font and cell.font.bold:
                    style_cmds.append(("FONTNAME", (t_col, tbl_row), (t_col, tbl_row), "Helvetica-Bold"))
            except Exception:
                pass

    table.setStyle(TableStyle(style_cmds))
    return table


class ExcelToPdfService:
    def __init__(self):
        pass

    async def analyze(self, request_id: str, filename: str) -> Dict[str, Any]:
        """
        Open the uploaded Excel file and extract the sheet names.
        Uses openpyxl (pure Python) — no Microsoft Excel required.
        """
        upload_dir = Paths.request_upload(request_id)
        input_path = upload_dir / filename

        if not input_path.exists():
            raise ValueError("File not found.")

        try:
            wb = openpyxl.load_workbook(str(input_path), read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            return {"filename": filename, "sheets": sheets}
        except Exception as e:
            logger.error(f"Failed to analyze Excel file {filename}: {str(e)}")
            raise ValueError(f"Failed to read Excel file: {str(e)}")

    async def process(self, request_id: str, filenames: List[str], config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert Excel files to PDF using openpyxl + reportlab.
        No Microsoft Excel installation required.
        """
        upload_dir = Paths.request_upload(request_id)
        output_dir = Paths.request_output(request_id)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = []

        # --- Page size & orientation ---
        page_size_key = config.get("page_size", "a4").lower()
        base_size = PAGE_SIZE_MAP.get(page_size_key, A4)
        orientation = config.get("orientation", "portrait").lower()
        if orientation == "landscape":
            page_size = landscape(base_size)
        else:
            page_size = base_size

        # --- Margins ---
        margins_preset = config.get("margins", "normal").lower()
        if margins_preset == "custom":
            lm = float(config.get("margin_left",  0.7)) * inch
            rm = float(config.get("margin_right", 0.7)) * inch
            tm = float(config.get("margin_top",   0.75)) * inch
            bm = float(config.get("margin_bottom",0.75)) * inch
        else:
            lm, rm, tm, bm = MARGIN_PRESETS.get(margins_preset, MARGIN_PRESETS["normal"])

        for filename in filenames:
            input_path = upload_dir / filename

            if not input_path.exists():
                results.append({"original_filename": filename, "status": "failed", "message": "File not found"})
                continue

            try:
                wb = openpyxl.load_workbook(str(input_path), data_only=True)
                selected_sheets = config.get("selected_sheets") or wb.sheetnames

                # Output filename
                custom_name = config.get("output_filename")
                if custom_name and len(filenames) == 1:
                    if not custom_name.lower().endswith(".pdf"):
                        custom_name += ".pdf"
                    output_filename = custom_name
                else:
                    output_filename = f"{input_path.stem}.pdf"

                output_path = output_dir / output_filename

                # Build PDF story
                doc = SimpleDocTemplate(
                    str(output_path),
                    pagesize=page_size,
                    leftMargin=lm,
                    rightMargin=rm,
                    topMargin=tm,
                    bottomMargin=bm,
                )

                styles = getSampleStyleSheet()
                story = []
                sheet_title_style = ParagraphStyle(
                    "SheetTitle",
                    parent=styles["Heading2"],
                    fontSize=10,
                    spaceAfter=6,
                )

                for i, sheet_name in enumerate(selected_sheets):
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]

                    if i > 0:
                        story.append(PageBreak())

                    story.append(Paragraph(f"Sheet: {sheet_name}", sheet_title_style))
                    story.append(Spacer(1, 6))

                    tbl = _sheet_to_table(ws, config)
                    story.append(tbl)

                wb.close()

                doc.build(story)

                if output_path.exists():
                    results.append({
                        "original_filename": filename,
                        "pdf_filename": output_filename,
                        "status": "success",
                    })
                else:
                    results.append({"original_filename": filename, "status": "failed", "message": "PDF not generated"})

            except Exception as e:
                logger.error(f"Error processing {filename}: {str(e)}")
                results.append({"original_filename": filename, "status": "failed", "message": str(e)})

        return {
            "success": any(r.get("status") == "success" for r in results),
            "request_id": request_id,
            "results": results,
        }


excel_to_pdf_service = ExcelToPdfService()
