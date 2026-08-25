"""
PDF to Excel conversion service.
Extracts tables, text, and paragraphs from each PDF page using PyMuPDF, writes them to an .xlsx file.
"""

import logging
import io
from pathlib import Path
from typing import Any, Dict, List, Tuple

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.paths import Paths

logger = logging.getLogger(__name__)


class PDFToExcelService:
    async def process(
        self,
        request_id: str,
        filename: str,
        config: Dict[str, Any] = None,
    ) -> Dict[str, Any]:
        """Convert a PDF to an Excel (.xlsx) file by extracting tables, text, and paragraphs."""
        if config is None:
            config = {}

        upload_dir = Paths.request_upload(request_id)
        output_dir = Paths.request_output(request_id)
        output_dir.mkdir(parents=True, exist_ok=True)

        pdf_path = upload_dir / filename
        if not pdf_path.exists():
            raise ValueError(f"File not found: {filename}")

        out_name = f"{pdf_path.stem}.xlsx"
        out_path = output_dir / out_name

        logger.info(f"Converting PDF to Excel: {pdf_path} -> {out_path}")

        try:
            doc = fitz.open(str(pdf_path))
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            title_fill = PatternFill("solid", fgColor="1E3A8A")
            title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2563EB")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            key_fill = PatternFill("solid", fgColor="F1F5F9")
            key_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
            alt_fill = PatternFill("solid", fgColor="F8FAFC")
            white_fill = PatternFill("solid", fgColor="FFFFFF")
            para_fill = PatternFill("solid", fgColor="EFF6FF")
            val_font = Font(name="Calibri", size=11, color="0F172A")
            section_font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")

            thin_border_side = Side(border_style="thin", color="CBD5E1")
            cell_border = Border(
                left=thin_border_side, right=thin_border_side,
                top=thin_border_side, bottom=thin_border_side,
            )
            align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            align_center = Alignment(horizontal="center", vertical="center")

            for page_num in range(len(doc)):
                page = doc[page_num]
                sheet_name = f"Page {page_num + 1}"
                ws = wb.create_sheet(title=sheet_name)
                ws.views.sheetView[0].showGridLines = True

                tables_data = []
                try:
                    tabs = page.find_tables()
                    for t in tabs:
                        ext = t.extract()
                        if ext and len(ext) >= 1:
                            tables_data.append(ext)
                except Exception:
                    pass

                table_rects = []
                if tables_data:
                    try:
                        for t in page.find_tables():
                            table_rects.append(t.bbox)
                    except Exception:
                        pass

                row_idx = 1

                if tables_data:
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                    title_cell = ws.cell(row=row_idx, column=1, value="TABLE DATA")
                    title_cell.fill = title_fill
                    title_cell.font = title_font
                    title_cell.alignment = align_center
                    title_cell.border = cell_border
                    ws.cell(row=row_idx, column=2).border = cell_border
                    ws.cell(row=row_idx, column=3).border = cell_border
                    ws.row_dimensions[row_idx].height = 28
                    row_idx += 1

                    for tbl_idx, tbl in enumerate(tables_data):
                        if len(tables_data) > 1:
                            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                            sec_cell = ws.cell(row=row_idx, column=1, value=f"Table {tbl_idx + 1}")
                            sec_cell.font = section_font
                            sec_cell.fill = PatternFill("solid", fgColor="E2E8F0")
                            sec_cell.alignment = align_left
                            sec_cell.border = cell_border
                            ws.cell(row=row_idx, column=2).border = cell_border
                            ws.cell(row=row_idx, column=3).border = cell_border
                            ws.row_dimensions[row_idx].height = 22
                            row_idx += 1

                        for r_i, row in enumerate(tbl):
                            for c_i, val in enumerate(row):
                                cell_val = str(val or "").strip()
                                cell = ws.cell(row=row_idx, column=c_i + 1, value=cell_val)
                                cell.border = cell_border
                                if r_i == 0:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = align_center
                                else:
                                    cell.fill = alt_fill if r_i % 2 == 0 else white_fill
                                    cell.font = val_font
                                    cell.alignment = align_left
                            row_idx += 1
                        row_idx += 1

                words = page.get_text("words")

                if not words or len(" ".join([w[4] for w in words]).strip()) < 10:
                    try:
                        import pytesseract
                        from PIL import Image
                        import shutil

                        tess_path = shutil.which("tesseract")
                        if not tess_path:
                            for tp in [
                                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                            ]:
                                if Path(tp).exists():
                                    pytesseract.pytesseract.tesseract_cmd = tp
                                    tess_path = tp
                                    break

                        if tess_path:
                            pix = page.get_pixmap(dpi=300)
                            img = Image.open(io.BytesIO(pix.tobytes("png")))
                            ocr_data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                            ocr_words = []
                            n_boxes = len(ocr_data.get("text", []))
                            for i in range(n_boxes):
                                txt = (ocr_data["text"][i] or "").strip()
                                if txt:
                                    left = ocr_data["left"][i]
                                    top = ocr_data["top"][i]
                                    width = ocr_data["width"][i]
                                    height = ocr_data["height"][i]
                                    x0 = left * (72.0 / 300.0)
                                    y0 = top * (72.0 / 300.0)
                                    x1 = (left + width) * (72.0 / 300.0)
                                    y1 = (top + height) * (72.0 / 300.0)
                                    ocr_words.append((x0, y0, x1, y1, txt))
                            if ocr_words:
                                words = ocr_words
                    except Exception:
                        pass

                text_words = []
                if words:
                    for w in words:
                        x0, y0, x1, y1, word_text = w[0], w[1], w[2], w[3], w[4]
                        word_rect = fitz.Rect(x0, y0, x1, y1)
                        in_table = False
                        for tr in table_rects:
                            if word_rect.intersects(tr):
                                in_table = True
                                break
                        if not in_table:
                            text_words.append((x0, y0, x1, y1, word_text))

                if text_words:
                    if tables_data:
                        row_idx += 1

                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
                    txt_title = ws.cell(row=row_idx, column=1, value="TEXT CONTENT")
                    txt_title.fill = title_fill
                    txt_title.font = title_font
                    txt_title.alignment = align_center
                    txt_title.border = cell_border
                    ws.cell(row=row_idx, column=2).border = cell_border
                    ws.cell(row=row_idx, column=3).border = cell_border
                    ws.row_dimensions[row_idx].height = 28
                    row_idx += 1

                    ws.cell(row=row_idx, column=1, value="Line #").font = header_font
                    ws.cell(row=row_idx, column=1).fill = header_fill
                    ws.cell(row=row_idx, column=1).alignment = align_center
                    ws.cell(row=row_idx, column=1).border = cell_border

                    ws.cell(row=row_idx, column=2, value="Content").font = header_font
                    ws.cell(row=row_idx, column=2).fill = header_fill
                    ws.cell(row=row_idx, column=2).alignment = align_center
                    ws.cell(row=row_idx, column=2).border = cell_border
                    ws.row_dimensions[row_idx].height = 22
                    row_idx += 1

                    lines_dict: Dict[float, List[Tuple[float, str]]] = {}
                    for w in text_words:
                        x0, y0, x1, y1, word_text = w[0], w[1], w[2], w[3], w[4]
                        line_key = round(y0 / 9.0) * 9.0
                        if line_key not in lines_dict:
                            lines_dict[line_key] = []
                        lines_dict[line_key].append((x0, word_text))

                    line_num = 1
                    for y_k in sorted(lines_dict.keys()):
                        line_words = sorted(lines_dict[y_k], key=lambda item: item[0])
                        full_line = " ".join([wt for _, wt in line_words]).strip()
                        if not full_line:
                            continue

                        ws.cell(row=row_idx, column=1, value=line_num).font = val_font
                        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="top")
                        ws.cell(row=row_idx, column=1).border = cell_border

                        ws.cell(row=row_idx, column=2, value=full_line).font = val_font
                        ws.cell(row=row_idx, column=2).alignment = align_left
                        ws.cell(row=row_idx, column=2).border = cell_border

                        row_idx += 1
                        line_num += 1

                    ws.column_dimensions["A"].width = 12
                    ws.column_dimensions["B"].width = 80
                    ws.column_dimensions["C"].width = 30

                elif not tables_data:
                    ws.cell(row=1, column=1, value="[Image/Scanned Document - No extractable text found]")
                    ws.cell(row=1, column=1).font = val_font

                if not tables_data and not text_words:
                    ws.column_dimensions["A"].width = 50
                else:
                    if not text_words:
                        ws.column_dimensions["A"].width = 34
                        ws.column_dimensions["B"].width = 52
                        ws.column_dimensions["C"].width = 30

            if not wb.sheetnames:
                ws = wb.create_sheet("Sheet1")
                ws["A1"] = "No extractable text found in PDF."

            doc.close()
            wb.save(str(out_path))

        except Exception as e:
            logger.error(f"PDF to Excel conversion failed: {e}", exc_info=True)
            raise ValueError(f"Failed to convert PDF to Excel: {e}")

        if not out_path.exists():
            raise ValueError("Conversion succeeded but output file is missing.")

        return {
            "success": True,
            "request_id": request_id,
            "output_filename": out_name,
            "original_filename": filename,
        }


pdf_to_excel_service = PDFToExcelService()
